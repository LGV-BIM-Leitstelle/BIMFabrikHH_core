from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


class ExcelStyler:
    def __init__(self, filename: object) -> object:
        self.filename = filename
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active
        self.font = Font(name="Aptos")
        self.font_bold = Font(name="Aptos", bold=True)
        self.cell_alignment = Alignment(horizontal="left", vertical="center")

        self.border = Border(
            top=Side(style=None),
            left=Side(style=None),
            right=Side(style=None),
            bottom=Side(style=None),
        )

    def apply_style(self):
        try:
            # Set font, border, and adjust column width
            for col in self.sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.3
                self.sheet.column_dimensions[column].width = adjusted_width

                # Apply font and border to all cells in the column
                for cell in col:
                    cell.font = self.font
                    cell.border = self.border
                    cell.alignment = self.cell_alignment

            # Apply gray background and bold font to header
            for cell in self.sheet[1]:
                cell.font = self.font_bold
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Add filter to header row
            self.sheet.auto_filter.ref = self.sheet.dimensions

            print("Styles applied successfully.")

        except Exception as e:
            print(f"An error occurred: {e}")

    def save(self):
        try:
            self.workbook.save(self.filename)
            print(f"File '{self.filename}' saved successfully.")
        except Exception as e:
            print(f"An error occurred while saving the file: {e}")
